import os
import argparse
import calendar
import datetime

from settings import *

import openpyxl
from openpyxl.styles import Border, Side


THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))


def get_report_date(file):
    year = datetime.datetime.now().year
    last_sheet_name = file.sheetnames[-1].capitalize()
    month = list(calendar.month_abbr).index(last_sheet_name)
    date = calendar.monthrange(year, month)[-1]
    report_date = '{}.{}.{}'.format(date, str(month).zfill(2), str(year)[2:])
    return report_date


def create_excel(input_path):
    input_file = openpyxl.load_workbook(input_path)
    report_date = get_report_date(input_file)
    wb = openpyxl.load_workbook('template.xlsx')
    ws = wb.active
    output_row = ws.max_row + 1
    for sheet in input_file:
        port = sheet.cell(*INPUT_FIRST_PORT).value
        for row_index in range(5, sheet.max_row):
            if sheet.cell(row_index, INPUT_PORT_COLUMN + 1).value is None:
                port = sheet.cell(row_index, INPUT_PORT_COLUMN).value
                continue
            if sheet.cell(row_index, INPUT_PORT_COLUMN).value == sheet.cell(*INPUT_TABLE_HEADER).value:
                continue
            parsing_data = {name: sheet.cell(row_index, column).value for name, column in INPUT_COLUMNS.items()}
            dynamic_attrs = {'report_date': report_date, 'port': port}
            create_output_row(ws, output_row, parsing_data, dynamic_attrs)
            apply_common_style(ws[output_row])
            output_row += 1
    dir, name = os.path.split(input_path)
    output_path = os.path.join(dir, '{} {}.xlsx'.format(name.split('.')[0], 'COMPILE'))
    wb.save(output_path)


def apply_common_style(row):
    for cell in row:
        cell.border = THIN_BORDER


def create_output_row(sheet, row_index, input_row_data, dynamic_attrs):
    for name, column in OUTPUT_COLUMNS.items():
        cell = sheet.cell(row_index, OUTPUT_COLUMNS[name])
        if name in STATIC_ATTRS:
            cell.value = STATIC_ATTRS[name]
        elif name in DYNAMIC_ATTRS:
            cell.value = dynamic_attrs[name]
        elif name in INPUT_COLUMNS:
            cell.value = input_row_data[name]
        cell.number_format = FORMAT.get(name, 'General')


def get_params():
    parser = argparse.ArgumentParser(add_help=True)
    parser.add_argument('input', help='excel file path')
    return parser.parse_args()


def remove_accessory_quotes(path):
    if path[0] == path[-1] == '"':
        path = path[1:-1]
    return path


def check_path(input_str):
    input = remove_accessory_quotes(input_str)
    if not os.path.isfile(input):
        raise FileExistsError('There is no file for this path ({})'.format(input))
    return input


if __name__ == "__main__":
    params = get_params()
    input_path = check_path(params.input)
    create_excel(input_path)
    print('Success!')



